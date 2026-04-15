#!/usr/bin/env python3
"""
看護師勤務表自動作成 - Streamlit アプリ
Excel / Google Sheets 両対応
"""
import streamlit as st
import pandas as pd
import calendar
from datetime import date
from io import BytesIO

import jpholiday

from shift_scheduler import (
    Staff, build_and_solve,
    D, N, A, O, R, V, E, L, ST, LD, SN,
    SHIFTS, DAY_SHIFTS, NIGHT_SHIFTS,
    TIER_A, TIER_AB, TIER_B, TIER_C, VALID_TIERS,
    _get_holidays_and_days_off, _write_one_sheet,
    SETTINGS_DEF, SETTINGS_KEYS,
    _parse_staff_list, _parse_requests, _parse_settings,
)

# ============================================================
# ページ設定
# ============================================================
st.set_page_config(
    page_title="勤務表作成ツール",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    /* ── 全体フォント ── */
    html, body, [class*="css"] { font-family: 'Noto Sans JP', 'Hiragino Kaku Gothic ProN', sans-serif; }

    /* ── メトリックカード ── */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #f8fbff 0%, #eef4fc 100%);
        border: 1px solid #d0e4f7;
        border-left: 4px solid #2563eb;
        padding: 14px 16px;
        border-radius: 10px;
        box-shadow: 0 1px 4px rgba(37,99,235,0.08);
    }
    div[data-testid="stMetric"] label { color: #4b6a8a; font-size: 0.78rem; font-weight: 600; letter-spacing: 0.03em; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { color: #1e3a5f; font-size: 1.6rem; font-weight: 700; }

    /* ── タブ ── */
    button[data-baseweb="tab"] { font-weight: 600; font-size: 0.88rem; }
    button[data-baseweb="tab"][aria-selected="true"] { color: #2563eb; border-bottom-color: #2563eb; }

    /* ── データフレーム ── */
    .stDataFrame { font-size: 12px; }
    .stDataFrame thead th { background: #1e3a5f !important; color: white !important; font-weight: 600; }

    /* ── ボタン ── */
    button[kind="primary"] { background: #2563eb; border-radius: 8px; font-weight: 600; }
    button[kind="primary"]:hover { background: #1d4ed8; }

    /* ── サイドバー ── */
    section[data-testid="stSidebar"] { background: #f0f5ff; }
    section[data-testid="stSidebar"] .stMarkdown h3 { color: #1e3a5f; font-size: 0.9rem; }

    /* ── 凡例バッジ ── */
    .shift-badge {
        display: inline-block; padding: 2px 10px; border-radius: 12px;
        font-size: 0.8rem; font-weight: 600; margin: 2px;
    }

    /* ── セクション区切り ── */
    hr { border: none; border-top: 2px solid #e2eaf5; margin: 1rem 0; }

    /* ── 警告・成功バナー強化 ── */
    div[data-testid="stAlert"] { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 定数・ヘルパー
# ============================================================
WEEKDAY_MAP = {"月": 0, "火": 1, "水": 2, "木": 3, "金": 4, "土": 5, "日": 6}
WEEKDAY_REV = {v: k for k, v in WEEKDAY_MAP.items()}

# ============================================================
# 日本看護協会 夜勤ガイドラインチェック
# ============================================================
def check_nursing_guidelines(schedule, names, tiers, r_dedicated, night_hours=16):
    """
    日本看護協会「夜勤・交代制勤務に関するガイドライン（2013年）」+
    労基法準拠チェック
    Args:
        night_hours: 1夜勤あたりの時間数（二交代=16h, 三交代=8h）
    """
    violations = []
    warnings = []
    ok_list = []

    for s in names:
        shifts = schedule[s]
        is_dedicated = r_dedicated.get(s, False)
        issues = []
        night_count = shifts.count(N)   # 通常夜勤
        sn_count = shifts.count(SN)     # 短夜勤
        total_night = night_count + sn_count

        # ── 1. 月の夜勤系回数（8回以内）────────────────────────────
        if not is_dedicated:
            if total_night > 8:
                issues.append({"rule": "夜勤回数", "level": "violation",
                                "detail": f"{total_night}回（N:{night_count}+準:{sn_count}）上限8回超過"})
            elif total_night == 8:
                issues.append({"rule": "夜勤回数", "level": "warning",
                                "detail": f"{total_night}回（N:{night_count}+準:{sn_count}）上限8回ちょうど"})

        # ── 2. 月夜勤時間72時間以内（日本看護協会ガイドライン）──
        # N×night_hours + SN×12h
        if not is_dedicated and night_hours > 0:
            total_hours = night_count * night_hours + sn_count * 12
            if total_hours > 72:
                issues.append({"rule": "72時間規制", "level": "violation",
                                "detail": f"月夜勤{total_hours}h（72h超過 / N:{night_count}×{night_hours}h+準:{sn_count}×12h）"})
            elif total_hours >= 64:
                issues.append({"rule": "72時間規制", "level": "warning",
                                "detail": f"月夜勤{total_hours}h（72h上限に近い）"})

        # ── 3. 連続夜勤（2回以内）───────────────────────────────
        i = 0; consec = 0; max_consec = 0
        while i < len(shifts):
            if shifts[i] == N:
                consec += 1; max_consec = max(max_consec, consec); i += 2
            else:
                consec = 0; i += 1
        if max_consec > 2:
            issues.append({"rule": "連続夜勤", "level": "violation",
                           "detail": f"最大{max_consec}連続（上限2回超過）"})

        # ── 4. 勤務間インターバル（11時間以上）──────────────────
        iv = [f"{d+2}日" for d in range(len(shifts)-1) if shifts[d]==A and shifts[d+1]==D]
        if iv:
            issues.append({"rule": "インターバル", "level": "violation",
                           "detail": f"明け翌日に日勤: {', '.join(iv)}（11時間未満の疑い）"})

        # ── 5. 夜勤翌日の明け確認 ──────────────────────────────
        for d in range(len(shifts)-1):
            if shifts[d] == N and shifts[d+1] != A:
                issues.append({"rule": "夜勤後の明け", "level": "violation",
                               "detail": f"{d+1}日: 夜勤翌日が明けでない（{shifts[d+1]}）"})

        # ── 集計 ─────────────────────────────────────────────
        total_hours = night_count * night_hours + sn_count * 12
        for item in issues:
            rec = {"名前": s, "Tier": tiers[s],
                   "夜勤回数": f"N:{night_count}+準:{sn_count}",
                   "月夜勤時間": f"{total_hours}h", **item}
            (violations if item["level"] == "violation" else warnings).append(rec)
        if not issues:
            ok_list.append(s)

    return violations, warnings, ok_list

def check_skill_pairing(schedule, names, tiers, num_days, year, month):
    """
    夜勤ペアのスキルバランスをチェックする。
    NG条件: 夜勤ペアが全員 C（新人のみ）の日がある。
    推奨: 各夜勤ペアにA/AB以上が1名以上含まれる。
    Returns:
        bad_days: list of dict（NG日の情報）
        warn_days: list of dict（注意日）
        ok_days: int
    """
    tier_rank = {TIER_A: 4, TIER_AB: 3, TIER_B: 2, TIER_C: 1}
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    bad_days = []
    warn_days = []
    ok_days = 0

    for d in range(num_days):
        night_staff = [s for s in names if schedule[s][d] in NIGHT_SHIFTS]
        if not night_staff:
            ok_days += 1
            continue
        wd = wdj[(fwd + d) % 7]
        day_label = f"{d+1}日({wd})"
        tier_list = [tiers.get(s, TIER_C) for s in night_staff]
        max_rank = max(tier_rank.get(t, 1) for t in tier_list)
        members_str = " + ".join(f"{s}({tiers.get(s,'?')})" for s in night_staff)

        if max_rank == 1:  # 全員C（新人のみ）
            bad_days.append({
                "日": day_label, "夜勤メンバー": members_str,
                "問題": "🚨 全員C（新人のみ）",
            })
        elif max_rank == 2:  # 最高がB
            warn_days.append({
                "日": day_label, "夜勤メンバー": members_str,
                "問題": "⚠ A/AB不在（Bが最高）",
            })
        else:
            ok_days += 1

    return bad_days, warn_days, ok_days


# ユニット種別ごとの配置基準定義
# (ratio_val, check_night, display_ratio, basis_note)
# ※ 各管理料にクラス（1〜Nまで）が存在する場合でも、
#    看護師配置比率はクラス共通。クラス差は医師在院要件・施設基準の厳格さ。
UNIT_STANDARDS = {
    "ICU（特定集中治療室管理料 1〜4）":    (2,  True,  "2:1",
        "特定集中治療室管理料1〜4 / 看護配置は全クラス共通 2:1 / "
        "クラス差: 専任常勤医師の在院要件・集中治療実績基準"),
    "HCU（ハイケアユニット入院医療管理料 1〜2）": (4, True, "4:1",
        "ハイケアユニット入院医療管理料1〜2 / 看護配置は全クラス共通 4:1 / "
        "クラス差: 専任医師配置・施設要件の厳格さ"),
    "NICU（新生児集中治療室管理料 1〜2）": (3,  True,  "3:1",
        "新生児集中治療室管理料1〜2 / 看護配置は全クラス共通 3:1 / "
        "クラス差: 治療実績・専従医師要件"),
    "GCU（新生児治療回復室入院医療管理料）": (6, True, "6:1",
        "新生児治療回復室入院医療管理料 / 6:1（クラス区分なし）"),
    "SCU（脳卒中ケアユニット入院医療管理料 1〜2）": (3, True, "3:1",
        "脳卒中ケアユニット入院医療管理料1〜2 / 看護配置は全クラス共通 3:1 / "
        "クラス差: 専任医師の常時在院か否か"),
    "PICU（小児特定集中治療室管理料 1〜2）": (2, True, "2:1",
        "小児特定集中治療室管理料1〜2 / 看護配置は全クラス共通 2:1 / "
        "クラス差: 専任常勤医師の在院要件"),
    "一般病棟（急性期一般入院料1） 7:1":   (7,  False, "7:1",
        "急性期一般入院料1 / 7:1（日勤帯）"),
    "一般病棟（急性期一般入院料4） 10:1":  (10, False, "10:1",
        "急性期一般入院料4 / 10:1（日勤帯）"),
    "一般病棟（急性期一般入院料6） 13:1":  (13, False, "13:1",
        "急性期一般入院料6 / 13:1（日勤帯）"),
    "一般病棟（地域一般入院基本料） 15:1": (15, False, "15:1",
        "地域一般入院基本料 / 15:1（日勤帯）"),
    "回復期リハビリテーション病棟":         (13, False, "13:1",
        "回復期リハビリテーション病棟入院料 / 13:1（日勤帯）"),
    "地域包括ケア病棟":                     (13, False, "13:1",
        "地域包括ケア病棟入院料 / 13:1（日勤帯）"),
}


def check_staffing_ratio(schedule, names, r_dedicated, r_weekly,
                          num_days, bed_count, ratio_val, year, month,
                          check_night=False):
    """
    日別の人員数が配置基準を満たすか確認する。
    Args:
        check_night: True の場合、夜勤帯（夜勤中＝N）も同一基準でチェック（ICU系）
    必要人数 = ceil(病床数 / 配置比)
    Returns:
        shortfalls: list of dict（不足日の情報）
        ok_days: 基準を満たす日数
        required: 必要人数
    """
    required = max(1, -(-bed_count // ratio_val))  # ceiling division
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    shortfalls = []
    ok_days = 0
    for d in range(num_days):
        wd = wdj[(fwd + d) % 7]
        day_staff = sum(1 for s in names if schedule[s][d] in DAY_SHIFTS)
        issues = []
        if day_staff < required:
            issues.append(f"日勤系{day_staff}人（必要{required}人）")
        if check_night:
            night_staff = sum(1 for s in names if schedule[s][d] in NIGHT_SHIFTS)
            if night_staff < required:
                issues.append(f"夜勤系{night_staff}人（必要{required}人）")
        if issues:
            shortfalls.append({
                "日": f"{d+1}日({wd})",
                "日勤人数": day_staff,
                **({"夜勤人数": sum(1 for s in names if schedule[s][d] == N)} if check_night else {}),
                "必要人数": required,
                "不足内容": " / ".join(issues),
            })
        else:
            ok_days += 1
    return shortfalls, ok_days, required


SHIFT_DISPLAY = {
    D: "日", N: "夜", A: "明", O: "休", R: "研", V: "暇",
    E: "早", L: "遅", ST: "短", LD: "長", SN: "準",
    "夜不": "夜不", "休暇": "休暇", "明休": "明休",
}
SHIFT_REVERSE = {
    "日": D, "夜": N, "明": A, "休": O, "研": R, "暇": V,
    "早": E, "遅": L, "短": ST, "長": LD, "準": SN,
    "夜不": "夜不", "休暇": "休暇", "明休": "明休",
}
# シフト説明（ツールチップ用）
SHIFT_DESC = {
    D: "日勤 (8:00〜17:00)",
    N: "夜勤 (16:45〜翌9:00 / 16h)",
    A: "明け（夜勤明け）",
    O: "公休",
    R: "研修",
    V: "休暇",
    E: "早出 (7:00〜16:00 / 8h)",
    L: "遅出 (12:00〜21:00 / 8h)",
    ST: "時短 (8:45〜16:00 / 6.25h) 育児・介護",
    LD: "長日勤 (8:45〜21:00 / 12h)",
    SN: "短夜勤 (17:00〜翌5:00 / 12h)",
}


def _staff_to_df(staff_list):
    """Staffリスト → DataFrame"""
    rows = []
    for s in staff_list:
        wd_str = ""
        if s.work_days:
            wd_str = "".join(WEEKDAY_REV[d] for d in sorted(s.work_days))
        rows.append({
            "名前": s.name, "Tier": s.tier,
            "夜勤専従": s.dedicated,
            "時短": s.short_time,
            "週勤務": s.weekly_days,
            "前月末": s.prev_month or "",
            "夜勤Min": s.night_min,
            "夜勤Max": s.night_max,
            "連勤Max": s.consec_max,
            "勤務曜日": wd_str,
            "祝日不可": s.no_holiday,
        })
    return pd.DataFrame(rows) if rows else _default_staff()


def _reqs_to_df(reqs_dict, staff_list, num_days):
    """希望dict → DataFrame"""
    req_rows = []
    for name, rq in reqs_dict.items():
        row = {"名前": name}
        for d in range(1, num_days + 1):
            v = rq.get(d, "")
            row[str(d)] = SHIFT_DISPLAY.get(v, v) if v else ""
        req_rows.append(row)
    existing = {r["名前"] for r in req_rows}
    for s in staff_list:
        if s.name not in existing:
            row = {"名前": s.name}
            for d in range(1, num_days + 1):
                row[str(d)] = ""
            req_rows.append(row)
    return pd.DataFrame(req_rows)


_SETTINGS_WIDGET_MAP = {
    "year": "inp_year", "month": "inp_month",
    "min_day_staff": "inp_min_day", "night_staff_count": "inp_night_count",
    "max_night_regular": "inp_max_n_reg", "pref_night_regular": "inp_pref_n_reg",
    "max_night_dedicated": "inp_max_n_ded", "pref_night_dedicated": "inp_pref_n_ded",
    "max_consecutive": "inp_max_consec", "pref_consecutive": "inp_pref_consec",
    "solver_time_limit": "inp_time_limit",
}

def _apply_settings(gs_settings):
    """読み込んだ設定を _pending_* キーに保存（ウィジェット描画前に反映させるため）"""
    for src, dst in _SETTINGS_WIDGET_MAP.items():
        v = gs_settings.get(src)
        if v is not None:
            st.session_state[f"_pending_{dst}"] = int(v)
    po = gs_settings.get("public_off_override")
    if po is not None and po != "":
        st.session_state["_pending_inp_po_mode"] = "手動指定"
        st.session_state["_pending_inp_po_val"] = int(po)
    else:
        st.session_state["_pending_inp_po_mode"] = "自動（土日祝）"


def _default_staff():
    return pd.DataFrame({
        "名前": ["スタッフA", "スタッフB", "スタッフC", "スタッフD", "スタッフE"],
        "Tier": ["A", "A", "AB", "C", "C"],
        "夜勤専従": [False, False, False, False, False],
        "時短": [False, False, False, False, False],
        "週勤務": [None, None, None, None, None],
        "前月末": ["", "", "", "", ""],
        "夜勤Min": [None, None, None, None, None],
        "夜勤Max": [None, None, None, None, None],
        "連勤Max": [None, None, None, None, None],
        "勤務曜日": ["", "", "", "", ""],
        "祝日不可": [False, False, False, False, False],
    })


# ============================================================
# 様式9帳票生成（看護職員夜勤・交代制勤務に関する実態調査）
# ============================================================
def _generate_youshiki9_excel(schedule, names, tiers, r_dedicated, r_weekly,
                               year, month, night_hours=16,
                               facility_name="", ward_name=""):
    """
    日本看護協会「様式9」帳票を生成。
    夜勤・交代制勤務に関する実態調査 月次提出用フォーム。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "様式9"

    thin = Side(style="thin")
    medium = Side(style="medium")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_bold = Border(left=medium, right=medium, top=medium, bottom=medium)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    sub_fill = PatternFill("solid", fgColor="D9E2F3")
    sub_font = Font(bold=True, size=9)
    warn_fill = PatternFill("solid", fgColor="F4CCCC")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")

    shift_system_label = "二交代制" if night_hours == 16 else "三交代制"

    # 列幅設定
    col_widths = {"A": 4, "B": 16, "C": 8, "D": 8, "E": 10, "F": 10,
                  "G": 10, "H": 10, "I": 10, "J": 14, "K": 10}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── タイトル ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "看護職員夜勤・交代制勤務に関する実態調査（様式9）"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")

    # ── 施設情報 ──
    ws.merge_cells("A2:C2")
    ws["A2"] = "調査年月"
    ws["A2"].font = sub_font
    ws.merge_cells("D2:E2")
    ws["D2"] = f"{year}年{month}月"
    ws["D2"].font = Font(bold=True, size=11)

    ws.merge_cells("A3:C3")
    ws["A3"] = "施設名"
    ws["A3"].font = sub_font
    ws.merge_cells("D3:G3")
    ws["D3"] = facility_name or "（施設名）"

    ws.merge_cells("A4:C4")
    ws["A4"] = "病棟・部署名"
    ws["A4"].font = sub_font
    ws.merge_cells("D4:G4")
    ws["D4"] = ward_name or "（病棟・部署名）"

    ws.merge_cells("H2:I2")
    ws["H2"] = "交代制区分"
    ws["H2"].font = sub_font
    ws.merge_cells("J2:K2")
    ws["J2"] = shift_system_label
    ws["J2"].font = Font(bold=True)

    ws.merge_cells("H3:I3")
    ws["H3"] = "夜勤時間/回"
    ws["H3"].font = sub_font
    ws.merge_cells("J3:K3")
    ws["J3"] = f"{night_hours}時間"
    ws["J3"].font = Font(bold=True)

    ws.merge_cells("H4:I4")
    ws["H4"] = "72時間規制 上限"
    ws["H4"].font = sub_font
    ws.merge_cells("J4:K4")
    ws["J4"] = f"{72 // night_hours}回/月"
    ws["J4"].font = Font(bold=True)

    # ── ヘッダー行 (row 6) ──
    headers = ["No", "氏名", "Tier", "雇用区分", "夜勤専従",
               "月夜勤回数", "月夜勤時間(h)", "72h規制", "連続夜勤(最大)",
               "インターバル違反", "判定"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    ws.row_dimensions[6].height = 28

    # ── データ行 ──
    def _max_consecutive_nights(shifts):
        i = 0; consec = 0; mx = 0
        while i < len(shifts):
            if shifts[i] == N:
                consec += 1; mx = max(mx, consec); i += 2
            else:
                consec = 0; i += 1
        return mx

    def _has_interval_violation(shifts):
        return any(shifts[d] == A and shifts[d+1] == D for d in range(len(shifts)-1))

    total_night_count = 0
    total_night_hours = 0
    violation_count = 0

    for idx, s in enumerate(names, 1):
        shifts = schedule[s]
        is_dedicated = r_dedicated.get(s, False)
        is_parttime = r_weekly.get(s) is not None
        night_cnt = shifts.count(N)
        total_h = night_cnt * night_hours
        max_consec = _max_consecutive_nights(shifts)
        iv_violation = _has_interval_violation(shifts)

        # 72h規制判定
        if is_dedicated:
            h72_label = "専従（対象外）"
            h72_ok = True
        elif total_h > 72:
            h72_label = f"🚨 {total_h}h 超過"
            h72_ok = False
        elif total_h >= 64:
            h72_label = f"⚠ {total_h}h"
            h72_ok = True
        else:
            h72_label = f"✅ {total_h}h"
            h72_ok = True

        # インターバル
        iv_label = "🚨 あり" if iv_violation else "✅ なし"

        # 総合判定
        is_violation = (not is_dedicated and total_h > 72) or iv_violation or max_consec > 2
        judgment = "🚨 要対応" if is_violation else "✅ 適合"

        row_data = [
            idx, s, tiers.get(s, "—"),
            "パート" if is_parttime else "常勤",
            "専従" if is_dedicated else "非専従",
            night_cnt, total_h, h72_label,
            max_consec, iv_label, judgment
        ]
        r = 6 + idx
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center")
            if c == 2:
                cell.alignment = Alignment(horizontal="left")
            if is_violation:
                cell.fill = warn_fill
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFF")

        if not is_dedicated:
            total_night_count += night_cnt
            total_night_hours += total_h
        if is_violation:
            violation_count += 1

    # ── 集計行 ──
    summary_row = 6 + len(names) + 1
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    ws[f"A{summary_row}"] = "集計"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"A{summary_row}"].fill = sub_fill

    ws.cell(row=summary_row, column=5, value=f"対象: {len(names)}名").font = Font(size=9)
    ws.cell(row=summary_row, column=6, value=f"合計: {total_night_count}回").font = Font(size=9, bold=True)
    ws.cell(row=summary_row, column=7, value=f"合計: {total_night_hours}h").font = Font(size=9, bold=True)
    ws.cell(row=summary_row, column=11,
            value=f"🚨 {violation_count}件" if violation_count else "✅ 全員適合")
    ws.cell(row=summary_row, column=11).font = Font(bold=True,
            color="CC0000" if violation_count else "006600")

    for c in range(1, 12):
        ws.cell(row=summary_row, column=c).border = bdr
        ws.cell(row=summary_row, column=c).fill = sub_fill

    # ── 注釈 ──
    note_row = summary_row + 2
    ws.merge_cells(f"A{note_row}:K{note_row}")
    ws[f"A{note_row}"] = (
        "【根拠】日本看護協会「夜勤・交代制勤務に関するガイドライン」(2013年) ／ "
        "夜勤回数: 月8回以内 ／ 72時間規制: 月夜勤時間72時間以内 ／ "
        "連続夜勤: 2サイクル以内 ／ 勤務間隔: 11時間以上"
    )
    ws[f"A{note_row}"].font = Font(size=8, color="888888", italic=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# Excelテンプレート生成
# ============================================================
def _generate_template_excel(year, month, num_staff=20):
    """入力用Excelテンプレートを生成"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    num_days = calendar.monthrange(year, month)[1]
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    # スタッフ情報エリア: 薄緑
    staff_hdr_fill = PatternFill("solid", fgColor="548235")
    staff_hdr_font = Font(bold=True, color="FFFFFF", size=10)
    staff_cell_fill = PatternFill("solid", fgColor="E2EFDA")
    # 勤務希望エリア: 青
    req_hdr_font = Font(bold=True, color="FFFFFF", size=9)
    req_cell_fill = PatternFill("solid", fgColor="D6E4F0")

    # --- Sheet 1: 設定 ---
    ws_s = wb.active
    ws_s.title = "設定"
    ws_s.column_dimensions["A"].width = 22
    ws_s.column_dimensions["B"].width = 14
    ws_s.column_dimensions["C"].width = 35
    # ヘッダー
    for c, txt in enumerate(["項目", "値", "説明"], 1):
        cell = ws_s.cell(row=3, column=c, value=txt)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    ws_s.cell(row=1, column=1, value="勤務表設定").font = Font(bold=True, size=14)
    # データ
    for i, (label, default, desc) in enumerate(SETTINGS_DEF):
        r = 4 + i
        ws_s.cell(row=r, column=1, value=label).border = bdr
        ws_s.cell(row=r, column=2, value=default).border = bdr
        ws_s.cell(row=r, column=3, value=desc).border = bdr
        ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)

    # コンプライアンス・運用条件（参考情報 — アプリ側のサイドバーで設定）
    comp_start = 4 + len(SETTINGS_DEF) + 2
    ws_s.cell(row=comp_start, column=1, value="コンプライアンス・運用条件（参考）").font = Font(
        bold=True, size=12, color="C00000")
    ws_s.cell(row=comp_start + 1, column=1, value="※以下はアプリのサイドバーで設定します（Excel上では参考情報）").font = Font(
        color="888888", size=9)
    comp_items = [
        ("72時間規制", "strict / soft / none",
         "strict=必ず準拠, soft=なるべく準拠(ペナルティ), none=チェックのみ"),
        ("長日勤→翌日短夜勤", "strict / soft / none",
         "LD(12h)の翌日にSN(12h)を入れるか"),
        ("長日勤の連続禁止", "strict / soft / none",
         "LD(12h)を2日連続で割り当てるか"),
        ("夜勤時間数", "16",
         "1夜勤あたりの時間数（二交代=16h）"),
        ("ユニット種別", "ICU",
         "ICU/HCU/NICU/GCU/SCU/PICU/一般病棟等"),
        ("病床数", "10",
         "人員配置基準の計算に使用"),
    ]
    for c, txt in enumerate(["項目", "設定値", "説明"], 1):
        cell = ws_s.cell(row=comp_start + 2, column=c, value=txt)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    for i, (label, default, desc) in enumerate(comp_items):
        r = comp_start + 3 + i
        ws_s.cell(row=r, column=1, value=label).border = bdr
        ws_s.cell(row=r, column=2, value=default).border = bdr
        ws_s.cell(row=r, column=2).font = Font(color="666666")
        ws_s.cell(row=r, column=3, value=desc).border = bdr
        ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)

    # --- Sheet 2: スタッフ・勤務希望（統合シート） ---
    ws_c = wb.create_sheet("スタッフ・勤務希望")
    ws_c.cell(row=1, column=1, value=f"スタッフ・勤務希望 — {year}年{month}月").font = Font(bold=True, size=14)
    ws_c.cell(row=2, column=1, value="緑エリア: スタッフ情報 ／ 青エリア: 勤務希望（日/夜/準/早/遅/長/短/休/研/夜不/休暇/明休）").font = Font(color="888888", size=9)

    # --- ヘッダー構築 ---
    staff_headers = ["名前", "Tier", "夜勤専従", "時短", "週勤務", "前月末",
                     "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可"]
    n_staff_cols = len(staff_headers)

    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    first_wd = date(year, month, 1).weekday()
    holidays = {d.day for d, _ in jpholiday.month_holidays(year, month)}

    # スタッフ情報ヘッダー (row 4) — 緑系
    for c, txt in enumerate(staff_headers, 1):
        cell = ws_c.cell(row=4, column=c, value=txt)
        cell.fill = staff_hdr_fill; cell.font = staff_hdr_font; cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    # グループラベル (row 3)
    ws_c.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_staff_cols)
    grp1 = ws_c.cell(row=3, column=1, value="👤 スタッフ情報")
    grp1.font = Font(bold=True, size=11, color="548235")
    grp1.alignment = Alignment(horizontal="center")
    grp1.fill = PatternFill("solid", fgColor="E2EFDA")

    # 勤務希望ヘッダー (row 3-4) — 青系
    day_start_col = n_staff_cols + 1
    ws_c.merge_cells(start_row=3, start_column=day_start_col,
                     end_row=3, end_column=day_start_col + num_days - 1)
    grp2 = ws_c.cell(row=3, column=day_start_col, value="📝 勤務希望")
    grp2.font = Font(bold=True, size=11, color="1F4E79")
    grp2.alignment = Alignment(horizontal="center")
    grp2.fill = PatternFill("solid", fgColor="D6E4F0")

    for d in range(1, num_days + 1):
        col = day_start_col + d - 1
        wd_name = weekdays_jp[(first_wd + d - 1) % 7]
        cell = ws_c.cell(row=4, column=col, value=f"{d}({wd_name})")
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr
        wd_idx = (first_wd + d - 1) % 7
        if d in holidays:
            cell.fill = PatternFill("solid", fgColor="F4CCCC")
            cell.font = Font(bold=True, color="CC0000", size=9)
        elif wd_idx >= 5:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
            cell.font = Font(bold=True, color="1F4E79", size=9)
        else:
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = req_hdr_font

    # --- 列幅 ---
    ws_c.column_dimensions["A"].width = 14
    for i in range(1, n_staff_cols):
        ws_c.column_dimensions[get_column_letter(i + 1)].width = 10
    for d in range(1, num_days + 1):
        ws_c.column_dimensions[get_column_letter(day_start_col + d - 1)].width = 7

    # --- サンプルデータ行（3行） + 空行（num_staff - 3行） ---
    samples = [
        ["山田太郎", "A", "", "", "", "", "", "", "", "", ""],
        ["佐藤花子", "AB", "", "", "", "", "", "", "", "", ""],
        ["鈴木一郎", "C", "", "", "3", "", "", "", "", "月水金", ""],
    ]
    total_rows = max(num_staff, len(samples))
    for i in range(total_rows):
        r = 5 + i
        if i < len(samples):
            row_data = samples[i]
        else:
            row_data = [f"スタッフ{i + 1}"] + [""] * (n_staff_cols - 1)
        # スタッフ情報セル（薄緑背景）
        for c, val in enumerate(row_data, 1):
            cell = ws_c.cell(row=r, column=c, value=val if val != "" else None)
            cell.border = bdr
            cell.fill = staff_cell_fill
        # 勤務希望セル（薄青背景・空欄）
        for d in range(1, num_days + 1):
            cell = ws_c.cell(row=r, column=day_start_col + d - 1)
            cell.border = bdr
            cell.fill = req_cell_fill

    # --- 凡例エリア（データの下に余白を空けて配置） ---
    legend_start_row = 5 + total_rows + 2
    # Tier定義
    ws_c.cell(row=legend_start_row, column=1, value="📖 Tier定義").font = Font(bold=True, size=11, color="548235")
    tier_defs = [
        ("A", "ベテラン・リーダー格（夜勤リーダー）"),
        ("AB", "中堅・リーダー代行可"),
        ("B", "一人立ち済み・通常メンバー"),
        ("C", "新人・経験浅い（A/AB/Bとペアに）"),
    ]
    for i, (tier, desc) in enumerate(tier_defs):
        ws_c.cell(row=legend_start_row + 1 + i, column=1, value=tier).font = Font(bold=True)
        ws_c.cell(row=legend_start_row + 1 + i, column=2, value=desc).font = Font(color="555555", size=9)

    # カラム説明
    ws_c.cell(row=legend_start_row + 6, column=1, value="📖 カラム説明").font = Font(bold=True, size=11, color="548235")
    col_defs = [
        ("夜勤専従", "ONで夜勤/明け/休のみのパターン"),
        ("時短", "ONで時短(ST)/早出/遅出/休のみ"),
        ("週勤務", "パートの週勤務日数（空欄=フルタイム）"),
        ("前月末", "前月末の勤務状態（夜/明）"),
        ("夜勤Min/Max", "個別の夜勤回数制限（空欄=全体設定）"),
        ("連勤Max", "最大連続勤務日数（空欄=全体設定）"),
        ("勤務曜日", "特定曜日のみ勤務（例:月水金）"),
        ("祝日不可", "ONで祝日に勤務を入れない"),
    ]
    for i, (col_name, desc) in enumerate(col_defs):
        ws_c.cell(row=legend_start_row + 7 + i, column=1, value=col_name).font = Font(bold=True, size=9)
        ws_c.cell(row=legend_start_row + 7 + i, column=2, value=desc).font = Font(color="555555", size=9)

    # シフト種別
    shift_legend_col = 5
    ws_c.cell(row=legend_start_row, column=shift_legend_col, value="📖 シフト種別").font = Font(bold=True, size=11, color="1F4E79")
    shift_legend = [
        ("日", "日勤 (8:00〜17:00)"), ("夜", "夜勤 (16:45〜翌9:00 / 16h)"),
        ("準", "短夜勤 (17:00〜翌5:00 / 12h)"), ("早", "早出 (7:00〜16:00)"),
        ("遅", "遅出 (12:00〜21:00)"), ("長", "長日勤 (8:45〜21:00 / 12h)"),
        ("短", "時短 (8:45〜16:00 / 6.25h)"), ("休", "公休"),
        ("研", "研修"), ("夜不", "この日は夜勤不可"),
        ("休暇", "有給休暇"), ("明休", "明け＋翌日も休み"),
    ]
    for i, (sym, desc) in enumerate(shift_legend):
        ws_c.cell(row=legend_start_row + 1 + i, column=shift_legend_col, value=sym).font = Font(bold=True, size=10)
        ws_c.cell(row=legend_start_row + 1 + i, column=shift_legend_col + 1, value=desc).font = Font(color="555555", size=9)

    # ウィンドウ枠の固定（ヘッダー + 名前列）
    ws_c.freeze_panes = "B5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# Excelアップロード解析
# ============================================================
def _parse_uploaded_excel(uploaded_file, year, month):
    """アップロードされたExcelを解析してスタッフ・希望・設定を返す"""
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)
    num_days = calendar.monthrange(year, month)[1]

    # 設定シート
    gs_settings = {}
    if "設定" in wb.sheetnames:
        ws = wb["設定"]
        rows = []
        for r in range(4, 4 + len(SETTINGS_KEYS)):
            row_vals = [ws.cell(row=r, column=c).value or "" for c in range(1, 4)]
            rows.append(row_vals)
        gs_settings = _parse_settings(rows)

    # --- 統合シート形式（新テンプレ） or 分離シート形式（旧テンプレ） ---
    staff_list = []
    reqs = {}
    n_staff_cols = 11  # 名前〜祝日不可

    if "スタッフ・勤務希望" in wb.sheetnames:
        # 新フォーマット（統合シート）
        ws = wb["スタッフ・勤務希望"]
        staff_rows = []
        req_rows = []
        for r in range(5, ws.max_row + 1):  # row 5 からデータ
            name_val = ws.cell(row=r, column=1).value
            if not name_val or not str(name_val).strip():
                continue
            # スタッフ情報（列 1〜11）
            s_vals = [ws.cell(row=r, column=c).value or "" for c in range(1, n_staff_cols + 2)]
            staff_rows.append(s_vals)
            # 勤務希望（列 12〜）
            day_start = n_staff_cols + 1
            r_vals = [str(name_val).strip()]
            for d in range(1, num_days + 1):
                v = ws.cell(row=r, column=day_start + d - 1).value or ""
                r_vals.append(v)
            req_rows.append(r_vals)
        if staff_rows:
            staff_list = _parse_staff_list(staff_rows)
        staff_names = [s.name for s in staff_list]
        if req_rows:
            reqs = _parse_requests(req_rows, staff_names, num_days)
    else:
        # 旧フォーマット（分離シート）
        if "スタッフ一覧" in wb.sheetnames:
            ws = wb["スタッフ一覧"]
            staff_rows = []
            for r in range(2, ws.max_row + 1):
                row_vals = [ws.cell(row=r, column=c).value or "" for c in range(1, 13)]
                if str(row_vals[0]).strip():
                    staff_rows.append(row_vals)
            staff_list = _parse_staff_list(staff_rows)
        if "勤務希望" in wb.sheetnames:
            ws = wb["勤務希望"]
            staff_names = [s.name for s in staff_list]
            req_rows = []
            for r in range(5, ws.max_row + 1):
                row_vals = [ws.cell(row=r, column=c).value or "" for c in range(1, num_days + 2)]
                if str(row_vals[0]).strip():
                    req_rows.append(row_vals)
            reqs = _parse_requests(req_rows, staff_names, num_days)

    return staff_list, reqs, gs_settings


# ============================================================
# セッション初期化
# ============================================================
if "staff_df" not in st.session_state:
    st.session_state.staff_df = _default_staff()
if "requests_df" not in st.session_state:
    st.session_state.requests_df = None
if "results" not in st.session_state:
    st.session_state.results = None
if "data_loaded" not in st.session_state:
    st.session_state.data_loaded = False

# ============================================================
# pending設定をウィジェットキーに反映（ウィジェット描画前に実行）
# ============================================================
for _src, _dst in _SETTINGS_WIDGET_MAP.items():
    _pk = f"_pending_{_dst}"
    if _pk in st.session_state:
        st.session_state[_dst] = st.session_state.pop(_pk)
for _pk, _wk in [("_pending_inp_po_mode", "inp_po_mode"), ("_pending_inp_po_val", "inp_po_val")]:
    if _pk in st.session_state:
        st.session_state[_wk] = st.session_state.pop(_pk)

# ============================================================
# サイドバー: 設定
# ============================================================
st.sidebar.title("勤務表設定")

col_y, col_m = st.sidebar.columns(2)
year = col_y.number_input("対象年", 2024, 2030, 2026, key="inp_year")
month = col_m.number_input("対象月", 1, 12, 5, key="inp_month")

num_days = calendar.monthrange(year, month)[1]
holidays_auto, weekends_auto, auto_public_off = _get_holidays_and_days_off(year, month)

st.sidebar.markdown("---")
st.sidebar.subheader("勤務条件")

public_off_mode = st.sidebar.radio("公休日数", ["自動（土日祝）", "手動指定"],
                                    horizontal=True, key="inp_po_mode")
if public_off_mode == "自動（土日祝）":
    public_off = auto_public_off
    st.sidebar.info(f"公休 {public_off}日（土日{len(weekends_auto)} + 祝{len(holidays_auto)}）")
    po_override = None
else:
    public_off = st.sidebar.number_input("公休日数", 0, num_days, auto_public_off, key="inp_po_val")
    po_override = public_off

min_day = st.sidebar.number_input("日勤最低人数", 1, 20, 5, key="inp_min_day")
night_count = st.sidebar.number_input("夜勤人数/日", 1, 5, 2, key="inp_night_count")

st.sidebar.markdown("---")
st.sidebar.subheader("夜勤設定")
col1, col2 = st.sidebar.columns(2)
max_n_reg = col1.number_input("通常 上限", 1, 15, 5, key="inp_max_n_reg")
pref_n_reg = col2.number_input("通常 推奨", 1, 15, 4, key="inp_pref_n_reg")
col3, col4 = st.sidebar.columns(2)
max_n_ded = col3.number_input("専従 上限", 1, 20, 10, key="inp_max_n_ded")
pref_n_ded = col4.number_input("専従 推奨", 1, 20, 9, key="inp_pref_n_ded")

st.sidebar.markdown("---")
st.sidebar.subheader("連勤設定")
col5, col6 = st.sidebar.columns(2)
max_consec = col5.number_input("最大連勤", 1, 10, 5, key="inp_max_consec")
pref_consec = col6.number_input("推奨連勤", 1, 10, 4, key="inp_pref_consec")

time_limit = st.sidebar.number_input("計算時間上限（秒）", 10, 600, 120, key="inp_time_limit")
num_patterns = st.sidebar.number_input("生成パターン数", 1, 5, 3, key="inp_num_patterns")

st.sidebar.markdown("---")
st.sidebar.subheader("⚖ コンプライアンス設定")
shift_system = st.sidebar.radio("交代制", ["二交代（16h夜勤）", "三交代（8h夜勤）"],
                                 horizontal=False, key="inp_shift_system")
night_hours = 16 if shift_system.startswith("二") else 8

_72h_limit = 72 // night_hours
night_72h_mode_label = st.sidebar.radio(
    f"72時間規制（≦{_72h_limit}回/月）",
    ["🚫 必ず準拠（ハード制約）", "⚠ なるべく準拠（ソフト制約）", "✅ 許容（チェックのみ）"],
    index=2,
    key="inp_72h_mode",
)
_72h_mode_map = {
    "🚫 必ず準拠（ハード制約）": "strict",
    "⚠ なるべく準拠（ソフト制約）": "soft",
    "✅ 許容（チェックのみ）": "none",
}
night_72h_mode = _72h_mode_map[night_72h_mode_label]
if night_72h_mode == "strict":
    st.sidebar.caption(f"✅ {night_hours}h×夜勤回数 を必ず72h以内に収めます（上限{_72h_limit}回）")
elif night_72h_mode == "soft":
    st.sidebar.caption(f"⚠ 72h超過を最小化しますが、達成できない場合は許容します")
else:
    st.sidebar.caption(f"✅ 制約なし（結果画面でのみ警告表示）")

st.sidebar.markdown("---")
st.sidebar.subheader("📋 運用条件ルール")

_rule_options = ["🚫 必ず準拠", "⚠ なるべく準拠", "✅ 許容"]
_rule_map = {"🚫 必ず準拠": "strict", "⚠ なるべく準拠": "soft", "✅ 許容": "none"}

ld_sn_label = st.sidebar.radio(
    "長日勤→翌日短夜勤",
    _rule_options, index=0,
    key="inp_ld_sn",
    help="長日勤(12h)の翌日に短夜勤(12h)を入れるか"
)
ld_consec_label = st.sidebar.radio(
    "長日勤の連続禁止",
    _rule_options, index=1,
    key="inp_ld_consec",
    help="長日勤(12h)を2日連続で割り当てるか"
)

op_rules = {
    "ld_sn": _rule_map[ld_sn_label],
    "ld_consecutive": _rule_map[ld_consec_label],
}

st.sidebar.markdown("---")
st.sidebar.subheader("🏥 人員配置基準")
unit_type = st.sidebar.selectbox(
    "ユニット種別",
    list(UNIT_STANDARDS.keys()),
    index=0,  # デフォルト: ICU
    key="inp_unit_type",
)
_ratio_val, _check_night, _ratio_label, _basis_label = UNIT_STANDARDS[unit_type]
bed_count = st.sidebar.number_input(
    "病床数（床）", min_value=1, max_value=200, value=10,
    key="inp_bed_count"
)
_required_nurses = max(1, -(-bed_count // _ratio_val))
_scope = "日勤・夜勤（24時間）" if _check_night else "日勤帯"
st.sidebar.caption(
    f"基準: {_ratio_label} ／ {_scope}\n"
    f"必要人数: {bed_count}床 ÷ {_ratio_val} = **{_required_nurses}人以上**"
)
st.sidebar.caption(f"根拠: {_basis_label}")
nurse_ratio = _ratio_label  # 表示用

settings = {
    "year": year, "month": month,
    "public_off_override": po_override,
    "min_day_staff": min_day, "night_staff_count": night_count,
    "max_night_regular": max_n_reg, "pref_night_regular": pref_n_reg,
    "max_night_dedicated": max_n_ded, "pref_night_dedicated": pref_n_ded,
    "max_consecutive": max_consec, "pref_consecutive": pref_consec,
    "solver_time_limit": time_limit, "holidays": "",
}

# ============================================================
# メインエリア
# ============================================================
st.title("🏥 勤務表自動作成ツール")

tab0, tab1, tab3, tab4, tab5 = st.tabs(
    ["📂 データ入力", "📋 スタッフ・勤務希望", "⚡ 生成", "📊 結果", "🏠 ダッシュボード"])

# ============================================================
# Tab 0: データ入力（テンプレ出力 / アップロード / スプシ）
# ============================================================
with tab0:
    st.subheader("データの入力方法を選んでください")

    col_left, col_right = st.columns(2)

    # --- 左カラム: テンプレートダウンロード ---
    with col_left:
        st.markdown("### 📥 テンプレートをダウンロード")
        st.markdown("""
        1. スタッフ人数を入力してテンプレートをダウンロード
        2. **スタッフ情報・勤務希望** を記入
        3. 右の「データ読み込み」で読み込み
        """)
        tmpl_staff_count = st.number_input(
            "テンプレートのスタッフ人数", min_value=1, max_value=100,
            value=15, step=1, key="tmpl_staff_count")
        template_bytes = _generate_template_excel(year, month, num_staff=tmpl_staff_count)
        st.download_button(
            label=f"📄 Excelテンプレート（{year}年{month}月・{tmpl_staff_count}人）",
            data=template_bytes,
            file_name=f"勤務表テンプレート_{year}_{month:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        # --- Googleスプレッドシートへテンプレート出力 ---
        if st.button("📊 Googleスプレッドシートにテンプレート作成", use_container_width=True, key="btn_gsheet_tmpl"):
            try:
                from shift_scheduler import _get_gsheet_client
                gc = _get_gsheet_client()
                title = f"勤務表テンプレート_{year}_{month:02d}"
                sh = gc.create(title)
                # --- 設定シート ---
                ws_s = sh.sheet1
                ws_s.update_title("設定")
                s_rows = [["項目", "値", "説明"]]
                for label, default, desc in SETTINGS_DEF:
                    s_rows.append([label, default, desc])
                ws_s.update(s_rows, "A1")
                # --- スタッフ一覧シート ---
                ws_st = sh.add_worksheet("スタッフ一覧", rows=30, cols=15)
                st_rows = [["名前", "Tier", "夜勤専従", "時短", "週勤務", "前月末",
                            "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可",
                            "", "Tier定義", "説明"]]
                st_rows.append(["山田太郎", "A", "", "", "", "", "", "", "", "", "",
                                "", "A", "ベテラン・リーダー格"])
                st_rows.append(["佐藤花子", "AB", "", "", "", "", "", "", "", "", "",
                                "", "AB", "中堅・リーダー代行可"])
                st_rows.append(["鈴木一郎", "C", "", "", "3", "", "", "", "", "月水金", "",
                                "", "B", "一人立ち済み"])
                st_rows.append(["", "", "", "", "", "", "", "", "", "", "",
                                "", "C", "新人・経験浅い"])
                ws_st.update(st_rows, "A1")
                # --- 勤務希望シート ---
                _num_days = calendar.monthrange(year, month)[1]
                ws_r = sh.add_worksheet("勤務希望", rows=30, cols=_num_days + 40)
                _wdj = ["月", "火", "水", "木", "金", "土", "日"]
                _fwd = date(year, month, 1).weekday()
                r_hdr = ["名前"] + [f"{d}({_wdj[(_fwd+d-1)%7]})" for d in range(1, _num_days+1)]
                # 凡例列
                _ref_col = _num_days + 2
                r_hdr += [""] + ["記号", "説明"]
                ws_r.update([r_hdr], "A1")
                legend = [
                    ["日", "日勤(8:00-17:00)"], ["夜", "夜勤(16:45-翌9:00)"],
                    ["準", "短夜勤(17:00-翌5:00)"], ["早", "早出(7:00-16:00)"],
                    ["遅", "遅出(12:00-21:00)"], ["長", "長日勤(8:45-21:00)"],
                    ["短", "時短(8:45-16:00)"], ["休", "公休"],
                    ["研", "研修"], ["夜不", "夜勤不可"], ["休暇", "有給休暇"],
                    ["明休", "明け＋翌日休み"],
                ]
                import gspread.utils as _gu
                _start = _gu.rowcol_to_a1(2, _num_days + 3)
                ws_r.update([[s, d] for s, d in legend], _start)

                url = sh.url
                st.success(f"✅ Googleスプレッドシートを作成しました")
                st.markdown(f"[📊 スプレッドシートを開く]({url})")
            except ImportError:
                st.error("gspread/google-authが必要です: pip install gspread google-auth")
            except Exception as e:
                st.error(f"作成エラー: {e}\n\n認証設定を確認してください（.streamlit/secrets.toml または credentials.json）")

    # --- 右カラム: データ読み込み ---
    with col_right:
        st.markdown("### 📤 データを読み込む")
        import_method = st.radio("読み込み方法", ["Excel ファイル", "Google スプレッドシート"],
                                  horizontal=True, key="import_method")

        # ローカルデフォルトファイルが存在する場合に表示
        import os as _os
        _default_excel = _os.path.join(_os.path.dirname(__file__), "勤務表_入力.xlsx")
        if _os.path.exists(_default_excel):
            if st.button("📁 ローカルファイルを読み込む（勤務表_入力.xlsx）",
                         use_container_width=True, key="btn_load_local"):
                try:
                    with st.spinner("読み込み中..."):
                        from shift_scheduler import load_input as _load_local
                        staff_list, reqs, gs_settings = _load_local()
                    if gs_settings:
                        _apply_settings(gs_settings)
                    st.session_state.staff_df = _staff_to_df(staff_list)
                    st.session_state.requests_df = _reqs_to_df(reqs, staff_list, num_days)
                    st.session_state.data_loaded = True
                    st.success(f"✅ {len(staff_list)}人のスタッフを読み込みました")
                    st.rerun()
                except Exception as e:
                    st.error(f"読み込みエラー: {e}")
            st.divider()

        if import_method == "Excel ファイル":
            uploaded = st.file_uploader("Excelファイルを選択", type=["xlsx", "xlsm"],
                                         key="excel_upload")
            if uploaded is not None:
                if st.button("📤 Excelから読み込み", type="primary", use_container_width=True,
                             key="btn_load_excel"):
                    try:
                        with st.spinner("読み込み中..."):
                            staff_list, reqs, gs_settings = _parse_uploaded_excel(
                                uploaded, year, month)
                        if gs_settings:
                            _apply_settings(gs_settings)
                        st.session_state.staff_df = _staff_to_df(staff_list)
                        st.session_state.requests_df = _reqs_to_df(reqs, staff_list, num_days)
                        st.session_state.data_loaded = True
                        st.success(f"✅ {len(staff_list)}人のスタッフと希望を読み込みました")
                        st.rerun()
                    except Exception as e:
                        st.error(f"読み込みエラー: {e}")

        else:  # Google スプレッドシート
            gsheet_id = st.text_input(
                "スプレッドシートID or URL",
                value="1mezi6NHOQZj0VR_TzmRb9UdkFGrFiWg7qJmvOrJopgA",
                key="gsheet_input")
            if st.button("📤 スプレッドシートから読み込み", type="primary",
                         use_container_width=True, key="btn_load_gsheet"):
                try:
                    with st.spinner("読み込み中..."):
                        from shift_scheduler import load_gsheet as _load_gs
                        staff_list, reqs_dict, gs_settings, _ = _load_gs(gsheet_id)
                    if gs_settings:
                        _apply_settings(gs_settings)
                    st.session_state.staff_df = _staff_to_df(staff_list)
                    st.session_state.requests_df = _reqs_to_df(reqs_dict, staff_list, num_days)
                    st.session_state.data_loaded = True
                    st.success(f"✅ {len(staff_list)}人のスタッフと希望を読み込みました")
                    st.rerun()
                except Exception as e:
                    st.error(f"読み込みエラー: {e}")

    # 読み込み状況
    st.markdown("---")
    if st.session_state.data_loaded:
        n_staff = len(st.session_state.staff_df.dropna(subset=["名前"]))
        n_staff = len([n for n in st.session_state.staff_df["名前"].dropna() if str(n).strip()])
        st.success(f"✅ データ読み込み済み（{n_staff}人）→ 「スタッフ・勤務希望」タブで確認・編集できます")
    else:
        st.info("💡 テンプレートをDLして記入 → アップロード、またはスプレッドシートから読み込んでください")

# ============================================================
# Tab 1: スタッフ・勤務希望（統合タブ）
# ============================================================
with tab1:
    st.subheader(f"スタッフ・勤務希望 — {year}年{month}月（{num_days}日間）")

    # --- スタッフ枠数設定 ---
    current_count = len(st.session_state.staff_df.dropna(subset=["名前"]))
    col_cnt, col_btn, col_spacer = st.columns([1, 1, 2])
    with col_cnt:
        target_count = st.number_input("スタッフ人数", min_value=1, max_value=100,
                                        value=max(current_count, 5), step=1, key="staff_total_count")
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button(f"✅ {target_count}人枠に確定", key="btn_set_staff_count"):
            current_df = st.session_state.staff_df
            valid_rows = current_df.dropna(subset=["名前"])
            valid_rows = valid_rows[valid_rows["名前"].str.strip() != ""]
            existing = len(valid_rows)
            if target_count > existing:
                new_rows = []
                for i in range(target_count - existing):
                    new_rows.append({
                        "名前": f"スタッフ{existing + i + 1}",
                        "Tier": "C", "夜勤専従": False, "時短": False,
                        "週勤務": None, "前月末": "", "夜勤Min": None,
                        "夜勤Max": None, "連勤Max": None, "勤務曜日": "", "祝日不可": False,
                    })
                st.session_state.staff_df = pd.concat(
                    [valid_rows, pd.DataFrame(new_rows)], ignore_index=True)
            elif target_count < existing:
                st.session_state.staff_df = valid_rows.head(target_count).reset_index(drop=True)
            st.rerun()

    # --- 表示モード切替 ---
    view_mode = st.radio(
        "表示モード",
        ["👤 スタッフ情報", "📝 勤務希望", "👤+📝 すべて"],
        horizontal=True, index=2, key="view_mode",
    )

    # --- スタッフ情報カラム選択（表示/非表示） ---
    _all_staff_detail_cols = ["Tier", "夜勤専従", "時短", "週勤務", "前月末",
                              "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可"]
    if view_mode in ("👤 スタッフ情報", "👤+📝 すべて"):
        with st.expander("⚙ 表示カラム選択", expanded=False):
            _visible_staff_cols = st.multiselect(
                "表示するスタッフ情報カラム",
                _all_staff_detail_cols,
                default=_all_staff_detail_cols,
                key="staff_col_toggle",
            )
    else:
        _visible_staff_cols = _all_staff_detail_cols

    # --- 定義パネル（表示モードに応じて出し分け） ---
    if view_mode in ("👤 スタッフ情報", "👤+📝 すべて"):
        with st.expander("📖 Tier（スキルランク）の定義", expanded=False):
            st.markdown("""
| Tier | 説明 | 夜勤の役割 |
|------|------|-----------|
| **A** | ベテラン・リーダー格 | 夜勤リーダー（必ずA/ABが1人以上） |
| **AB** | 中堅・リーダー代行可 | Aが不在時にリーダー代行 |
| **B** | 一人立ち済み | 通常メンバー（B+Cだけのペアは回避） |
| **C** | 新人・経験浅い | 通常メンバー（必ずA/AB/Bとペアに） |

**各カラムの説明:**
- **夜勤専従**: ONにすると夜勤・明け・休のみの勤務パターンになります
- **時短**: ONにすると時短(ST)・早出(E)・遅出(L)・休のみ。夜勤/長日勤は入りません
- **週勤務**: パートタイムの場合に週あたりの勤務日数を指定（空欄=フルタイム）
- **前月末**: 前月最終日の勤務（夜=夜勤中/明=夜勤明け）→翌月初の制約に反映
- **夜勤Min/Max**: 個別の夜勤回数制限（空欄=全体設定に従う）
- **連勤Max**: 最大連続勤務日数（空欄=全体設定に従う）
- **勤務曜日**: パート等で特定曜日のみ勤務する場合（例: 月水金）
- **祝日不可**: ONにすると祝日に勤務を入れません
""")

    if view_mode in ("📝 勤務希望", "👤+📝 すべて"):
        with st.expander("📖 シフト種別の定義", expanded=False):
            st.markdown("""
| 記号 | 正式名 | 時間帯 | 備考 |
|------|--------|--------|------|
| **日** | 日勤 | 8:00〜17:00 (8h) | 標準勤務 |
| **夜** | 夜勤 | 16:45〜翌9:00 (16h) | 翌日は自動で「明け」 |
| **準** | 短夜勤 | 17:00〜翌5:00 (12h) | 夜勤の短縮版 |
| **早** | 早出 | 7:00〜16:00 (8h) | 日勤の早番 |
| **遅** | 遅出 | 12:00〜21:00 (8h) | 日勤の遅番 |
| **長** | 長日勤 | 8:45〜21:00 (12h) | 日勤の延長版 |
| **短** | 時短 | 8:45〜16:00 (6.25h) | 育児・介護向け |
| **休** | 公休 | — | 通常の休日 |
| **研** | 研修 | — | 研修日（勤務扱い） |
| **夜不** | 夜勤不可 | — | この日は夜勤に入れないで |
| **休暇** | 有給休暇 | — | 休暇申請済み |
| **明休** | 明け＋休 | — | 明けの翌日も休みに |

**注意:** 夜勤専従スタッフは「早・遅」を選択しないでください（自動で除外されます）
""")

    # --- 日付ヘッダー準備 ---
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    header_map = {}
    for d in range(1, num_days + 1):
        wd = wdj[(fwd + d - 1) % 7]
        is_hol = d in holidays_auto
        is_we = d in weekends_auto
        suffix = "祝" if is_hol else ("★" if is_we else "")
        header_map[str(d)] = f"{d}({wd}{suffix})"

    # --- スタッフ名の同期（staff_df → requests_df） ---
    staff_names = [n for n in st.session_state.staff_df["名前"].dropna().tolist() if str(n).strip()]
    if st.session_state.requests_df is None or set(st.session_state.requests_df["名前"].tolist()) != set(staff_names):
        if st.session_state.requests_df is not None:
            old = st.session_state.requests_df.set_index("名前")
            rows = []
            for name in staff_names:
                if name in old.index:
                    rows.append(old.loc[name].to_dict() | {"名前": name})
                else:
                    row = {"名前": name}
                    for d in range(1, num_days + 1):
                        row[str(d)] = ""
                    rows.append(row)
            st.session_state.requests_df = pd.DataFrame(rows)
        else:
            rows = []
            for name in staff_names:
                row = {"名前": name}
                for d in range(1, num_days + 1):
                    row[str(d)] = ""
                rows.append(row)
            st.session_state.requests_df = pd.DataFrame(rows)

    # --- 統合DataFrame構築 ---
    _staff_cols = ["Tier", "夜勤専従", "時短", "週勤務", "前月末",
                   "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可"]
    _day_cols = [str(d) for d in range(1, num_days + 1)]

    # staff_df と requests_df を名前で結合
    _sdf = st.session_state.staff_df.copy()
    _rdf = st.session_state.requests_df.copy()
    # 名前の順序はstaff_dfを優先
    combined_df = _sdf.set_index("名前").reindex(staff_names)
    _rdf_indexed = _rdf.set_index("名前").reindex(staff_names)
    for col in _day_cols:
        if col in _rdf_indexed.columns:
            combined_df[col] = _rdf_indexed[col].values
        else:
            combined_df[col] = ""
    combined_df = combined_df.reset_index()

    # --- 表示モードに応じたカラム選択（トグル反映） ---
    if view_mode == "👤 スタッフ情報":
        show_cols = ["名前"] + _visible_staff_cols
    elif view_mode == "📝 勤務希望":
        show_cols = ["名前"] + _day_cols
    else:
        show_cols = ["名前"] + _visible_staff_cols + _day_cols

    # --- column_config 構築 ---
    _col_config_defs = {
        "Tier": st.column_config.SelectboxColumn("Tier", options=["A", "AB", "B", "C"], width="small"),
        "夜勤専従": st.column_config.CheckboxColumn("夜勤専従", width="small"),
        "時短": st.column_config.CheckboxColumn("時短", width="small"),
        "週勤務": st.column_config.NumberColumn("週勤務", min_value=1, max_value=7, step=1, width="small"),
        "前月末": st.column_config.SelectboxColumn("前月末", options=["", "夜", "明"], width="small"),
        "夜勤Min": st.column_config.NumberColumn("夜勤Min", min_value=0, max_value=15, step=1, width="small"),
        "夜勤Max": st.column_config.NumberColumn("夜勤Max", min_value=0, max_value=15, step=1, width="small"),
        "連勤Max": st.column_config.NumberColumn("連勤Max", min_value=1, max_value=10, step=1, width="small"),
        "勤務曜日": st.column_config.TextColumn("勤務曜日", width="small", help="例: 月火木"),
        "祝日不可": st.column_config.CheckboxColumn("祝日不可", width="small"),
    }
    col_config = {}
    col_config["名前"] = st.column_config.TextColumn("名前", width="medium")
    # 選択されたスタッフ情報カラムのみ追加
    if view_mode != "📝 勤務希望":
        for c in _visible_staff_cols:
            if c in _col_config_defs:
                col_config[c] = _col_config_defs[c]
    # 勤務希望カラム
    if view_mode != "👤 スタッフ情報":
        shift_options = ["", "日", "夜", "準", "早", "遅", "長", "短", "休", "研", "夜不", "休暇", "明休"]
        for d in range(1, num_days + 1):
            col_config[str(d)] = st.column_config.SelectboxColumn(
                header_map[str(d)], options=shift_options, width="small")

    # --- data_editor（統合） ---
    edited_combined = st.data_editor(
        combined_df[show_cols],
        num_rows="dynamic" if view_mode != "📝 勤務希望" else "fixed",
        column_config=col_config,
        use_container_width=True,
        key="combined_editor",
        hide_index=True,
    )

    # --- 編集結果を元のDFに書き戻す ---
    # スタッフ情報
    _staff_out_cols = [c for c in (["名前"] + _staff_cols) if c in edited_combined.columns]
    edited_staff = edited_combined[_staff_out_cols].copy()
    # 勤務希望モードで名前しか表示していない場合は元のstaff_dfを保持
    if view_mode == "📝 勤務希望":
        # staff_dfの名前だけ更新し、属性はそのまま
        edited_staff = st.session_state.staff_df.copy()
    st.session_state.staff_df = edited_staff

    # 勤務希望
    _req_out_cols = [c for c in (["名前"] + _day_cols) if c in edited_combined.columns]
    if len(_req_out_cols) > 1:  # 名前+日付列がある
        edited_reqs = edited_combined[_req_out_cols].copy()
        st.session_state.requests_df = edited_reqs
    else:
        edited_reqs = st.session_state.requests_df

    if holidays_auto:
        hol_names = {d.day: name for d, name in jpholiday.month_holidays(year, month)}
        hol_str = "、".join(f"{d}日({hol_names.get(d, '祝')})" for d in sorted(holidays_auto))
        st.info(f"🗓 {year}年{month}月の祝日: {hol_str}")

# ============================================================
# Tab 3: 生成
# ============================================================
with tab3:
    st.subheader("勤務表生成")

    col_a, col_b, col_c = st.columns(3)
    valid_staff = edited_staff.dropna(subset=["名前"])
    valid_staff = valid_staff[valid_staff["名前"].str.strip() != ""]
    ft_count = len(valid_staff[valid_staff["週勤務"].isna()])
    pt_count = len(valid_staff[valid_staff["週勤務"].notna()])
    ded_count = len(valid_staff[valid_staff["夜勤専従"] == True])

    col_a.metric("スタッフ数", f"{len(valid_staff)}人")
    col_b.metric("フルタイム / パート", f"{ft_count} / {pt_count}")
    col_c.metric("夜勤専従", f"{ded_count}人")

    st.markdown("---")

    if st.button("🚀 勤務表を生成", type="primary", use_container_width=True):
        staff_list = []
        for _, row in valid_staff.iterrows():
            name = str(row["名前"]).strip()
            tier = str(row["Tier"]).strip() if pd.notna(row["Tier"]) else "C"
            if tier not in VALID_TIERS:
                st.warning(f"⚠ {name}: Tier '{tier}' 不正 → スキップ")
                continue
            ded = bool(row.get("夜勤専従", False))
            weekly = int(row["週勤務"]) if pd.notna(row.get("週勤務")) else None
            prev = str(row.get("前月末", "")).strip()
            if prev not in ("夜", "明", ""):
                prev = ""
            n_min = int(row["夜勤Min"]) if pd.notna(row.get("夜勤Min")) else None
            n_max = int(row["夜勤Max"]) if pd.notna(row.get("夜勤Max")) else None
            c_max = int(row["連勤Max"]) if pd.notna(row.get("連勤Max")) else None
            wd_str = str(row.get("勤務曜日", "")).strip()
            work_days = None
            if wd_str:
                work_days = set()
                for ch in wd_str:
                    if ch in WEEKDAY_MAP:
                        work_days.add(WEEKDAY_MAP[ch])
                if not work_days:
                    work_days = None
            no_hol = bool(row.get("祝日不可", False))
            short_t = bool(row.get("時短", False))
            staff_list.append(Staff(name, tier, ded, weekly, prev,
                                     n_min, n_max, c_max, work_days, no_hol, short_t))

        if not staff_list:
            st.error("スタッフが0人です")
        else:
            reqs_dict = {}
            dedicated_names = {s.name for s in staff_list if s.dedicated}
            for _, row in edited_reqs.iterrows():
                name = str(row["名前"]).strip()
                if not name:
                    continue
                rq = {}
                for d in range(1, num_days + 1):
                    val = str(row.get(str(d), "")).strip()
                    if val and val in SHIFT_REVERSE:
                        rq[d] = SHIFT_REVERSE[val]
                    elif val and val in (D, N, O, R):
                        rq[d] = val
                # 夜勤専従: 早出・遅出の希望を除外
                if name in dedicated_names:
                    dropped = [d for d, v in rq.items() if v in (E, L)]
                    if dropped:
                        st.warning(f"⚠ {name}（専従）: {len(dropped)}日分の早出/遅出希望を無視しました")
                        for d in dropped:
                            del rq[d]
                if rq:
                    reqs_dict[name] = rq

            with st.spinner(f"最適化計算中... （最大{time_limit}秒 × {num_patterns}パターン）"):
                import io, contextlib
                console = io.StringIO()
                with contextlib.redirect_stdout(console):
                    results = build_and_solve(staff_list, reqs_dict, settings,
                                              num_patterns=num_patterns,
                                              night_hours=night_hours,
                                              night_72h_mode=night_72h_mode,
                                              op_rules=op_rules)

            console_output = console.getvalue()

            if results:
                st.session_state.results = results
                st.session_state.console_output = console_output
                st.success(f"✅ {len(results)}パターン生成完了！「結果」タブで確認できます。")

                buf = BytesIO()
                from openpyxl import Workbook
                wb = Workbook()
                wb.remove(wb.active)
                for res in results:
                    pat = res.get("pattern_num", 1)
                    title = f"パターン{pat}" if len(results) > 1 else "勤務表"
                    _write_one_sheet(wb, res, title)
                wb.save(buf)
                buf.seek(0)
                st.session_state.excel_bytes = buf.getvalue()
            else:
                st.error("❌ 解なし（Infeasible）。設定・希望を見直してください。")
                with st.expander("ソルバーログ"):
                    st.code(console_output)

# ============================================================
# Tab 4: 結果表示
# ============================================================
with tab4:
    if st.session_state.results is None:
        st.info("「生成」タブで勤務表を作成してください。")
    else:
        results = st.session_state.results

        if len(results) > 1:
            pat_idx = st.selectbox("パターン選択",
                                    range(len(results)),
                                    format_func=lambda i: f"パターン {i+1}")
        else:
            pat_idx = 0

        result = results[pat_idx]
        schedule = result["schedule"]
        names = result["names"]
        tiers = result["tiers"]
        r_num_days = result["num_days"]
        r_year = result["year"]
        r_month = result["month"]
        r_holidays = result.get("holidays", set())
        r_weekends = result.get("weekends", set())
        r_public_off = result.get("public_off", 13)
        r_weekly = result.get("weekly", {})
        r_dedicated = result.get("dedicated", {})
        missed = result.get("missed_requests", {})

        st.subheader(f"📊 {r_year}年{r_month}月 勤務表 — パターン {pat_idx + 1}")

        # 夜勤系合計（N + SN）
        night_counts = {s: sum(schedule[s].count(t) for t in NIGHT_SHIFTS) for s in names}
        nc_regular = [v for s, v in night_counts.items()
                      if r_weekly.get(s) is None and not r_dedicated.get(s, False)]
        total_missed = sum(len(v) for v in missed.values())
        violations_pre, _, _ = check_nursing_guidelines(
            schedule, names, tiers, r_dedicated, night_hours)
        bad_pairs_pre, _, _ = check_skill_pairing(
            schedule, names, tiers, r_num_days, r_year, r_month)
        shortfalls_pre, _, _ = check_staffing_ratio(
            schedule, names, r_dedicated, r_weekly,
            r_num_days, bed_count, _ratio_val, r_year, r_month,
            check_night=_check_night)

        col1, col2, col3, col4, col5, col6 = st.columns(6)
        col1.metric("公休日数", f"{r_public_off}日")
        col2.metric("夜勤均等", f"{min(nc_regular)}〜{max(nc_regular)}回" if nc_regular else "—")
        col3.metric("未達希望", f"{total_missed}件" if total_missed else "0件 ✓",
                    delta="要確認" if total_missed else None, delta_color="inverse")
        col4.metric("ガイドライン", "✅ 適合" if not violations_pre else f"🚨 {len(violations_pre)}件",
                    delta="要対応" if violations_pre else None, delta_color="inverse")
        col5.metric("スキルペア", "✅ 適正" if not bad_pairs_pre else f"🚨 {len(bad_pairs_pre)}日",
                    delta="要対応" if bad_pairs_pre else None, delta_color="inverse")
        col6.metric("配置基準", "✅ 適合" if not shortfalls_pre else f"🚨 {len(shortfalls_pre)}日",
                    delta="要対応" if shortfalls_pre else None, delta_color="inverse")

        wdj = ["月", "火", "水", "木", "金", "土", "日"]
        fwd = date(r_year, r_month, 1).weekday()

        shift_colors = {
            D: "#FFFFFF",   # 日勤: 白
            N: "#4472C4",   # 夜勤: 青
            A: "#FFF2CC",   # 明け: 薄黄
            O: "#E2EFDA",   # 休み: 薄緑
            R: "#E8D5F5",   # 研修: 薄紫
            V: "#F4CCCC",   # 休暇: 薄赤
            E:  "#E6F3FF",  # 早出: 水色
            L:  "#FFF0E0",  # 遅出: 薄橙
            ST: "#F0FFF0",  # 時短: 薄緑系
            LD: "#FFF5CC",  # 長日勤: 薄黄緑
            SN: "#2E75B6",  # 短夜勤: 濃青
        }
        shift_text_colors = {N: "#FFFFFF", SN: "#FFFFFF"}

        day_cols = [f"{d+1}" for d in range(r_num_days)]
        table_data = []
        for s in names:
            row = {"名前": s, "Tier": tiers[s]}
            for d in range(r_num_days):
                row[day_cols[d]] = schedule[s][d]
            counts = {t: schedule[s].count(t) for t in SHIFTS}
            row["日"] = counts[D]
            row["夜"] = counts[N]
            row["準"] = counts[SN]
            row["明"] = counts[A]
            row["早"] = counts[E]
            row["遅"] = counts[L]
            row["長"] = counts[LD]
            row["短"] = counts[ST]
            row["休"] = counts[O]
            row["研"] = counts[R]
            row["暇"] = counts[V]
            row["公休"] = counts[O] + counts[V]
            table_data.append(row)

        df = pd.DataFrame(table_data)

        def color_shifts(val):
            if val in shift_colors:
                bg = shift_colors[val]
                fg = shift_text_colors.get(val, "#000000")
                return f"background-color: {bg}; color: {fg}; text-align: center; font-weight: {'bold' if val == N else 'normal'}"
            return "text-align: center"

        # ── シフト表示モード切替 ─────────────────────────────
        view_mode = st.radio("表示モード", ["👁 確認", "✏️ 手動編集"],
                             horizontal=True, key=f"view_mode_{pat_idx}")

        if view_mode == "👁 確認":
            styled = df.style.map(color_shifts, subset=day_cols)
            st.dataframe(styled, use_container_width=True, height=600, hide_index=True)
        else:
            st.caption("⚠ セルを直接編集できます。変更後「変更を保存」を押してください。")
            shift_options_edit = [D, N, SN, A, E, L, LD, ST, O, R, V]
            col_cfg_edit = {
                "名前": st.column_config.TextColumn("名前", disabled=True, width="small"),
                "Tier": st.column_config.TextColumn("Tier", disabled=True, width="small"),
            }
            for dc in day_cols:
                col_cfg_edit[dc] = st.column_config.SelectboxColumn(
                    dc, options=shift_options_edit, width="small")
            for col in ["日", "夜", "準", "明", "早", "遅", "長", "短", "休", "研", "暇", "公休"]:
                col_cfg_edit[col] = st.column_config.NumberColumn(col, disabled=True, width="small")

            edited_df = st.data_editor(
                df, column_config=col_cfg_edit,
                use_container_width=True, height=600, hide_index=True,
                key=f"manual_edit_{pat_idx}"
            )
            if st.button("💾 変更を保存", key=f"save_edit_{pat_idx}"):
                # edited_df の内容をresultsに反映
                new_schedule = {}
                for _, row in edited_df.iterrows():
                    sname = row["名前"]
                    new_schedule[sname] = [row[dc] for dc in day_cols]
                st.session_state.results[pat_idx]["schedule"] = new_schedule
                st.success("✅ 変更を保存しました。ページを再読み込みすると反映されます。")
                st.rerun()

        # ── 夜勤回数分布チャート ─────────────────────────────
        with st.expander("📊 スタッフ別 夜勤・勤務統計", expanded=False):
            stat_rows = []
            for s in names:
                sh = schedule[s]
                n_cnt = sh.count(N)
                sn_cnt = sh.count(SN)
                total_h = n_cnt * night_hours + sn_cnt * 12
                stat_rows.append({
                    "名前": s, "Tier": tiers[s],
                    "夜勤(N)": n_cnt,
                    "短夜勤(準)": sn_cnt,
                    "月夜勤時間": f"{total_h}h",
                    "72h": "🚨" if total_h > 72 else ("⚠" if total_h >= 64 else "✅"),
                    "早出": sh.count(E), "遅出": sh.count(L),
                    "長日勤": sh.count(LD), "時短": sh.count(ST),
                    "日勤": sh.count(D),
                    "休み": sh.count(O) + sh.count(V),
                })
            stat_df = pd.DataFrame(stat_rows)
            st.dataframe(stat_df, use_container_width=True, hide_index=True)

            st.markdown("**夜勤回数分布**")
            chart_data = pd.DataFrame({
                "夜勤回数": [night_counts[s] for s in names]
            }, index=names)
            st.bar_chart(chart_data, height=200)

        with st.expander("🌙 日別夜勤ペア", expanded=False):
            pair_data = []
            for d in range(r_num_days):
                wd = wdj[(fwd + d) % 7]
                hol = "祝" if (d+1) in r_holidays else ("★" if (d+1) in r_weekends else "")
                nn_n  = [s for s in names if schedule[s][d] == N]
                nn_sn = [s for s in names if schedule[s][d] == SN]
                members = (
                    [f"{s}({tiers[s]})" for s in nn_n] +
                    [f"{s}({tiers[s]})[準]" for s in nn_sn]
                )
                pair_data.append({
                    "日": f"{d+1}日({wd}{hol})",
                    "夜勤メンバー": " + ".join(members) if members else "—",
                })
            st.dataframe(pd.DataFrame(pair_data), use_container_width=True, hide_index=True)

        if missed:
            with st.expander("⚠ 未達希望", expanded=True):
                for s, ds in missed.items():
                    st.warning(f"{s}: {', '.join(f'{d}日' for d in ds)}")
        else:
            st.success("✓ 全希望達成")

        # ── 日本看護協会 + 労基法ガイドラインチェック ─────────
        violations, warnings_gl, ok_list = check_nursing_guidelines(
            schedule, names, tiers, r_dedicated, night_hours
        )
        gl_label = "✅ ガイドライン適合" if not violations else f"🚨 違反 {len(violations)}件"
        with st.expander(f"⚖ 夜勤ガイドライン & 労基法チェック — {gl_label}", expanded=bool(violations)):
            st.caption(f"根拠: 日本看護協会「夜勤・交代制勤務に関するガイドライン」(2013年) ／ 交代制: {shift_system}")
            gl_col1, gl_col2, gl_col3 = st.columns(3)
            gl_col1.metric("🚨 違反", f"{len(violations)}件",
                           delta="要対応" if violations else None, delta_color="inverse")
            gl_col2.metric("⚠ 注意", f"{len(warnings_gl)}件")
            gl_col3.metric("✅ 適合", f"{len(ok_list)}名")

            st.markdown("##### チェック項目")
            rules_info = [
                ("夜勤回数", "月8回以内（夜勤専従者は対象外）"),
                ("72時間規制", f"月夜勤時間72時間以内（{night_hours}h×回数）"),
                ("連続夜勤", "連続夜勤2サイクル以内"),
                ("インターバル", "勤務間隔11時間以上（明け翌日に日勤がないか）"),
                ("夜勤後の明け", "夜勤翌日は必ず明けシフト"),
            ]
            for rule, desc in rules_info:
                v_cnt = sum(1 for x in violations if x["rule"] == rule)
                w_cnt = sum(1 for x in warnings_gl if x["rule"] == rule)
                icon = "🚨" if v_cnt else ("⚠" if w_cnt else "✅")
                st.markdown(f"- {icon} **{rule}**: {desc}"
                            + (f" → **{v_cnt}件違反**" if v_cnt else "")
                            + (f" → {w_cnt}件注意" if w_cnt else ""))

            if violations:
                st.markdown("##### 🚨 違反一覧")
                vcols = ["名前", "Tier", "夜勤回数", "月夜勤時間", "rule", "detail"]
                st.dataframe(
                    pd.DataFrame(violations)[[c for c in vcols if c in pd.DataFrame(violations).columns]]
                    .rename(columns={"rule": "項目", "detail": "内容"}),
                    use_container_width=True, hide_index=True
                )

            if warnings_gl:
                st.markdown("##### ⚠ 注意一覧")
                st.dataframe(
                    pd.DataFrame(warnings_gl)[["名前", "Tier", "夜勤回数", "月夜勤時間", "rule", "detail"]]
                    .rename(columns={"rule": "項目", "detail": "内容"}),
                    use_container_width=True, hide_index=True
                )

            if ok_list:
                st.markdown(f"##### ✅ 全項目適合: {', '.join(ok_list)}")

        # ── スキルペアチェック ───────────────────────────────
        bad_pairs, warn_pairs, ok_pair_days = check_skill_pairing(
            schedule, names, tiers, r_num_days, r_year, r_month)
        pair_label = ("✅ 全日適正" if not bad_pairs and not warn_pairs
                      else f"🚨 NG {len(bad_pairs)}日" if bad_pairs
                      else f"⚠ 注意 {len(warn_pairs)}日")
        with st.expander(f"👥 夜勤スキルペアチェック — {pair_label}",
                         expanded=bool(bad_pairs)):
            st.caption("A/AB（ベテラン）が各夜勤ペアに1名以上含まれるかチェックします。")
            sp1, sp2, sp3 = st.columns(3)
            sp1.metric("🚨 全員新人（C）日", f"{len(bad_pairs)}日",
                       delta="要対応" if bad_pairs else None, delta_color="inverse")
            sp2.metric("⚠ A/AB不在日", f"{len(warn_pairs)}日")
            sp3.metric("✅ 適正ペア日", f"{ok_pair_days}日")

            tier_legend = "Tier: **A**=エキスパート / **AB**=上級 / **B**=中堅 / **C**=新人"
            st.caption(tier_legend)

            if bad_days := bad_pairs:
                st.markdown("##### 🚨 問題あり（新人のみ夜勤）")
                st.dataframe(pd.DataFrame(bad_days), use_container_width=True, hide_index=True)
            if warn_days := warn_pairs:
                st.markdown("##### ⚠ 注意（A/AB不在）")
                st.dataframe(pd.DataFrame(warn_days), use_container_width=True, hide_index=True)
            if not bad_pairs and not warn_pairs:
                st.success("全夜勤ペアにA/AB（ベテラン）が含まれています。")

        # ── 人員配置基準チェック ─────────────────────────────
        shortfalls, ok_days, req_nurses = check_staffing_ratio(
            schedule, names, r_dedicated, r_weekly,
            r_num_days, bed_count, _ratio_val, r_year, r_month,
            check_night=_check_night
        )
        ratio_label = "✅ 基準適合" if not shortfalls else f"🚨 不足 {len(shortfalls)}日"
        scope_label = "日勤・夜勤（24時間）" if _check_night else "日勤帯"
        with st.expander(
            f"🏥 人員配置基準チェック — {unit_type}（{nurse_ratio}） — {ratio_label}",
            expanded=bool(shortfalls)
        ):
            st.caption(f"根拠: {_basis_label}")
            st.caption(f"配置基準: {nurse_ratio} ／ 適用範囲: {scope_label} ／ 病床数: {bed_count}床 ／ 必要人数: {req_nurses}人以上")
            rc1, rc2, rc3 = st.columns(3)
            rc1.metric("🚨 不足日数", f"{len(shortfalls)}日",
                       delta="要対応" if shortfalls else None, delta_color="inverse")
            rc2.metric("✅ 基準達成", f"{ok_days}日")
            rc3.metric("必要人数", f"{req_nurses}人/{scope_label}")
            if shortfalls:
                st.dataframe(pd.DataFrame(shortfalls), use_container_width=True, hide_index=True)
            else:
                st.success(f"全{r_num_days}日、{scope_label} {req_nurses}人以上を確保しています。")

        with st.expander("📈 日別集計", expanded=False):
            summary_data = {"日付": [f"{d+1}" for d in range(r_num_days)]}
            for label, shift in [
                ("日勤", D), ("夜勤", N), ("短夜勤", SN), ("明け", A),
                ("早出", E), ("遅出", L), ("長日勤", LD), ("時短", ST),
                ("休み", O), ("研修", R),
            ]:
                summary_data[label] = [sum(1 for s in names if schedule[s][d] == shift)
                                        for d in range(r_num_days)]
            # 日勤系合計列を追加
            summary_data["日勤系計"] = [
                sum(1 for s in names if schedule[s][d] in DAY_SHIFTS)
                for d in range(r_num_days)
            ]
            st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)

        with st.expander("🔧 ソルバーログ"):
            st.code(st.session_state.get("console_output", ""))

        st.markdown("---")
        dl_col1, dl_col2 = st.columns(2)

        if "excel_bytes" in st.session_state:
            dl_col1.download_button(
                label="📥 勤務表 Excelダウンロード",
                data=st.session_state.excel_bytes,
                file_name=f"勤務表_{r_year}_{r_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        # ── 様式9帳票 ──
        with st.expander("📋 様式9帳票（夜勤・交代制勤務実態調査）を出力", expanded=False):
            st.caption("日本看護協会・都道府県看護協会への月次提出用フォームです。")
            f9_col1, f9_col2 = st.columns(2)
            f9_facility = f9_col1.text_input("施設名", placeholder="○○病院", key=f"f9_facility_{pat_idx}")
            f9_ward = f9_col2.text_input("病棟・部署名", placeholder="ICU・CCU", key=f"f9_ward_{pat_idx}")

            youshiki9_bytes = _generate_youshiki9_excel(
                schedule, names, tiers, r_dedicated, r_weekly,
                r_year, r_month, night_hours,
                facility_name=f9_facility, ward_name=f9_ward
            )
            dl_col2.download_button(
                label="📋 様式9帳票ダウンロード",
                data=youshiki9_bytes,
                file_name=f"様式9_{r_year}_{r_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_youshiki9_{pat_idx}",
            )

# ============================================================
# Tab 5: ダッシュボード
# ============================================================
with tab5:
    st.subheader("🏠 ダッシュボード — 月次KPI概要")

    if st.session_state.results is None:
        st.info("「生成」タブで勤務表を作成すると、ここにKPIが表示されます。")
    else:
        # 最新パターンの結果を使用
        _res = st.session_state.results[0]
        _schedule = _res["schedule"]
        _names = _res["names"]
        _tiers = _res["tiers"]
        _r_num_days = _res["num_days"]
        _r_year = _res["year"]
        _r_month = _res["month"]
        _r_dedicated = _res.get("dedicated", {})
        _r_weekly = _res.get("weekly", {})
        _r_public_off = _res.get("public_off", 0)
        _missed = _res.get("missed_requests", {})

        st.caption(f"{_r_year}年{_r_month}月 ／ スタッフ {len(_names)}名 ／ {_r_num_days}日間")

        # ── KPIカード ──
        _v, _w, _ok = check_nursing_guidelines(_schedule, _names, _tiers, _r_dedicated, night_hours)
        _bp, _wp, _okp = check_skill_pairing(_schedule, _names, _tiers, _r_num_days, _r_year, _r_month)
        _sf, _okd, _req = check_staffing_ratio(
            _schedule, _names, _r_dedicated, _r_weekly,
            _r_num_days, bed_count, _ratio_val, _r_year, _r_month,
            check_night=_check_night)
        _total_missed = sum(len(v) for v in _missed.values())
        _nc = {s: _schedule[s].count(N) for s in _names}
        _nc_reg = [v for s, v in _nc.items()
                   if _r_weekly.get(s) is None and not _r_dedicated.get(s, False)]

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("スタッフ数", f"{len(_names)}人",
                  delta=f"うち専従 {sum(1 for s in _names if _r_dedicated.get(s))}人")
        k2.metric("公休日数", f"{_r_public_off}日")
        k3.metric("夜勤回数（均等）",
                  f"{min(_nc_reg)}〜{max(_nc_reg)}回" if _nc_reg else "—")
        k4.metric("未達希望", f"{_total_missed}件" if _total_missed else "0件 ✓",
                  delta="要確認" if _total_missed else None, delta_color="inverse")

        st.markdown("---")
        k5, k6, k7, k8 = st.columns(4)
        k5.metric("ガイドライン違反", f"{len(_v)}件",
                  delta="🚨 要対応" if _v else "✅ 適合", delta_color="inverse" if _v else "off")
        k6.metric("スキルペアNG日", f"{len(_bp)}日",
                  delta="🚨 要対応" if _bp else "✅ 適正", delta_color="inverse" if _bp else "off")
        k7.metric("配置基準不足日", f"{len(_sf)}日",
                  delta="🚨 要対応" if _sf else "✅ 適合", delta_color="inverse" if _sf else "off")
        k8.metric("コンプライアンス",
                  "✅ 全適合" if not _v and not _bp and not _sf else "🚨 要対応",
                  delta=None)

        st.markdown("---")
        dash_col1, dash_col2 = st.columns(2)

        # ── Tier分布 ──
        with dash_col1:
            st.markdown("##### 👥 スタッフ Tier分布")
            tier_counts = {}
            for s in _names:
                t = _tiers.get(s, "C")
                tier_counts[t] = tier_counts.get(t, 0) + 1
            tier_df = pd.DataFrame([
                {"Tier": t, "人数": tier_counts.get(t, 0)}
                for t in ["A", "AB", "B", "C"]
            ])
            st.dataframe(tier_df, use_container_width=True, hide_index=True)
            st.bar_chart(tier_df.set_index("Tier"), height=180)

        # ── 夜勤回数分布 ──
        with dash_col2:
            st.markdown("##### 🌙 夜勤回数分布（通常スタッフ）")
            reg_names = [s for s in _names
                         if _r_weekly.get(s) is None and not _r_dedicated.get(s, False)]
            if reg_names:
                nc_df = pd.DataFrame({
                    "夜勤回数": [_nc[s] for s in reg_names]
                }, index=reg_names)
                st.bar_chart(nc_df, height=180)
            else:
                st.info("通常スタッフなし")

        st.markdown("---")
        # ── 日別シフト構成 ──
        st.markdown("##### 📅 日別 シフト人数推移")
        daily_df = pd.DataFrame({
            "日勤": [sum(1 for s in _names if _schedule[s][d] == D) for d in range(_r_num_days)],
            "夜勤": [sum(1 for s in _names if _schedule[s][d] == N) for d in range(_r_num_days)],
            "休み": [sum(1 for s in _names if _schedule[s][d] in (O, V)) for d in range(_r_num_days)],
        }, index=[f"{d+1}日" for d in range(_r_num_days)])
        st.area_chart(daily_df, height=220)

        # ── 問題サマリー ──
        issues_any = bool(_v or _bp or _sf or _total_missed)
        if issues_any:
            st.markdown("---")
            st.markdown("##### 🚨 要対応サマリー")
            if _v:
                st.error(f"ガイドライン違反: {len(_v)}件 → 「結果」タブの⚖ガイドラインチェックを確認")
            if _bp:
                st.error(f"新人のみ夜勤: {len(_bp)}日 → 「結果」タブの👥スキルペアチェックを確認")
            if _sf:
                st.warning(f"配置基準不足: {len(_sf)}日 → 「結果」タブの🏥人員配置基準チェックを確認")
            if _total_missed:
                st.warning(f"未達希望: {_total_missed}件 → 「結果」タブの⚠未達希望を確認")
        else:
            st.success("✅ 全チェック項目に問題ありません。勤務表は提出可能な状態です。")
